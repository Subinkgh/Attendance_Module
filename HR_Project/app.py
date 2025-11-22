
from flask import Flask, render_template, request, jsonify
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Utility: Normalize column names consistently
def normalize_columns(cols):
    return [str(col).strip().replace(" ", "").replace(".", "").lower() for col in cols]

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        emp_id = request.form.get('empId')
        report_type = request.form.get('reportType')
        month = request.form.get('month')
        try:
            df = pd.read_excel('data/attendance.xlsx', sheet_name='PunchReport', engine='openpyxl')
            df.columns = [col.strip() for col in df.columns]

            # Validate required columns
            required_cols = ['Card No', 'Date', 'MusterMark']
            for col in required_cols:
                if col not in df.columns:
                    return f"Missing required column: {col}"

            # Parse dates safely
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df = df.dropna(subset=['Date'])
            df['Month'] = df['Date'].dt.strftime('%Y-%m')

            filtered = df[(df['Card No'].astype(str) == emp_id) & (df['Month'] == month)]
            if filtered.empty:
                return "No data found for the given Employee ID and Month."

            summary = {
                'Total Days': len(filtered),
                'PP': (filtered['MusterMark'] == 'PP').sum(),
                'PA': (filtered['MusterMark'] == 'PA').sum(),
                'AP': (filtered['MusterMark'] == 'AP').sum(),
                'AA': (filtered['MusterMark'] == 'AA').sum(),
                'OO': (filtered['MusterMark'] == 'OO').sum(),
                'HH': (filtered['MusterMark'] == 'HH').sum(),
                'EL': (filtered['MusterMark'] == 'EL').sum(),
                'Half_EL': (filtered['MusterMark'] == '1/2EL').sum(),
                'Additional': filtered.get('Additional', pd.Series([0]*len(filtered))).sum()
            }
            summary['Present Days'] = summary['PP'] + 0.5 * (summary['PA'] + summary['AP'])
            summary['Absent Days'] = summary['AA'] + 0.5 * (summary['PA'] + summary['AP'])
            summary['Earned Leave (EL)'] = summary['EL'] + 0.5 * summary['Half_EL']
            employee_name = filtered.iloc[0].get('Employee Name', '')

            template = 'report.html' if report_type == 'Punch Report' else 'muster.html'
            return render_template(template, data=filtered.to_dict(orient='records'), emp_id=emp_id, emp_name=employee_name, month=month, summary=summary)
        except Exception as e:
            return "Error loading report. Please check the file format and data integrity."
    return render_template('form.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        os.makedirs('data', exist_ok=True)
        if 'excelFile' not in request.files:
            return jsonify(success=False, message="No file part")
        file = request.files['excelFile']
        if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify(success=False, message="Invalid file format. Only .xlsx and .xls allowed.")

        temp_path = os.path.join('data', 'uploaded_temp.xlsx')
        file.save(temp_path)

        try:
            wb = load_workbook(temp_path, data_only=True)
            punch_sheet = next((wb[s] for s in wb.sheetnames if s.startswith("Punch Report")), None)
            if not punch_sheet:
                os.remove(temp_path)
                return jsonify(success=False, message="No valid Punch Report sheet found.")

            # Dynamic header detection
            headers = [str(cell.value).strip() if cell.value else "" for cell in punch_sheet[4]]
            sub_headers = [str(cell.value).strip() if cell.value else "" for cell in punch_sheet[5]]

            data_rows = []
            for row in punch_sheet.iter_rows(min_row=7, values_only=True):
                if all(cell is None for cell in row):
                    continue
                row_dict = dict(zip(headers, row))
                row_dict["Punch-Timings First"] = sub_headers[0]
                row_dict["Punch-Timings Last"] = sub_headers[1]
                normalized_row = {str(k).strip().replace(" ", "").replace(".", "").lower(): v for k, v in row_dict.items()}
                data_rows.append(normalized_row)

            if not data_rows:
                os.remove(temp_path)
                return jsonify(success=False, message="No valid punch data found.")

            punch_dates = [pd.to_datetime(r.get("punchdate"), errors='coerce') for r in data_rows if r.get("punchdate")]
            punch_dates = [d for d in punch_dates if pd.notnull(d)]
            if not punch_dates:
                os.remove(temp_path)
                return jsonify(success=False, message="No valid PunchDate found.")

            month_year = punch_dates[0].strftime("%Y-%m")
            csv_file_path = f"data/{month_year}.csv"

            # Prepare DataFrame
            expected_cols = normalize_columns(["SrNo.", "Card No", "ECode", "Employee Name", "PunchDate", "Day", "Punch-Timings First", "Punch-Timings Last", "Shift", "MusterMark", "WorkTime", "ActualHours", "Late", "Early", "OverTime", "Remarks", "Additional"])
            if os.path.exists(csv_file_path):
                df_existing = pd.read_csv(csv_file_path)
                df_existing.columns = normalize_columns(df_existing.columns)
            else:
                df_existing = pd.DataFrame(columns=expected_cols)

            new_rows = []
            for row in data_rows:
                card_no = row.get("cardno")
                punch_date = row.get("punchdate")
                if not card_no or pd.isnull(punch_date):
                    continue

                key_match = (df_existing["cardno"] == card_no) & (df_existing["punchdate"] == punch_date)
                try:
                    overtime = float(row.get("overtime", 0)) if row.get("overtime") not in [None, "", " "] else 0
                except (ValueError, TypeError):
                    overtime = 0
                additional = round(overtime, 2)  # FIX: Removed *24 inflation

                new_row = {col: row.get(col, "") for col in expected_cols}
                new_row["additional"] = additional

                if not df_existing[key_match].empty:
                    for k, v in new_row.items():
                        df_existing.loc[key_match, k] = v
                else:
                    new_rows.append(new_row)

            if new_rows:
                df_existing = pd.concat([df_existing, pd.DataFrame(new_rows)], ignore_index=True)

            df_existing.to_csv(csv_file_path, index=False)
            os.remove(temp_path)
            return jsonify(success=True, message=f"Punch data saved to {month_year}.csv")
        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify(success=False, message="Error processing file. Please check the data format.")
    return render_template('upload.html')

if __name__ == '__main__':
    app.run(debug=True)
